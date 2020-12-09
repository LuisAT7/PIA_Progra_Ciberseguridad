#!/usr/bin/env python3


import os
import random
import sys
import re
import requests
from ExtraerData import Scraping
import logging
from pyhunter import PyHunter
from openpyxl import Workbook
import hashlib
import socket



def Sistea_Op():
    if os.name == "nt":
        operating_system = "windows"
    if os.name == "posix":
        operating_system = "posix"
    return operating_system
if Sistea_Op() == "posix":
    class colores:
        PURPLE = '\033[95m'
        CYAN = '\033[96m'
        DARKCYAN = '\033[36m'
        BLUE = '\033[94m'
        GREEN = '\033[92m'
        YELLOW = '\033[93m'
        RED = '\033[91m'
        BOLD = '\033[1m'
        UNDERL = '\033[4m'
        ENDC = '\033[0m'
        backBlack = '\033[40m'
        backRed = '\033[41m'
        backGreen = '\033[42m'
        backYellow = '\033[43m'
        backBlue = '\033[44m'
        backMagenta = '\033[45m'
        backCyan = '\033[46m'
        backWhite = '\033[47m'

        def disable(self):
            self.PURPLE = ''
            self.CYAN = ''
            self.BLUE = ''
            self.GREEN = ''
            self.YELLOW = ''
            self.RED = ''
            self.ENDC = ''
            self.BOLD = ''
            self.UNDERL = ''
            self.backBlack = ''
            self.backRed = ''
            self.backGreen = ''
            self.backYellow = ''
            self.backBlue = ''
            self.backMagenta = ''
            self.backCyan = ''
            self.backWhite = ''
            self.DARKCYAN = ''

# Para Windows
else:
    class colores:
        PURPLE = '\033[95m'
        CYAN = '\033[96m'
        DARKCYAN = '\033[36m'
        BLUE = '\033[94m'
        GREEN = '\033[92m'
        YELLOW = '\033[93m'
        RED = '\033[91m'
        BOLD = '\033[1m'
        UNDERL = '\033[4m'
        ENDC = '\033[0m'
        backBlack = '\033[40m'
        backRed = '\033[41m'
        backGreen = '\033[42m'
        backYellow = '\033[43m'
        backBlue = '\033[44m'
        backMagenta = '\033[45m'
        backCyan = '\033[46m'
        backWhite = '\033[47m'

        def disable(self):
            self.PURPLE = ''
            self.CYAN = ''
            self.BLUE = ''
            self.GREEN = ''
            self.YELLOW = ''
            self.RED = ''
            self.ENDC = ''
            self.BOLD = ''
            self.UNDERL = ''
            self.backBlack = ''
            self.backRed = ''
            self.backGreen = ''
            self.backYellow = ''
            self.backBlue = ''
            self.backMagenta = ''
            self.backCyan = ''
            self.backWhite = ''
            self.DARKCYAN = ''

def sacar_correos():
        url = input('Ingresa el url de donde deseas obtener los correos: ')

        response = requests.get(url)
        if response.status_code != 200:
            exit()
        regExMail = r"[a-z0-9\.\-+_]+@[a-z0-9\.\-+_]+\.[a-z]+"
        new_emails = set(re.findall(regExMail, response.text, re.I))
        for i in new_emails:
            print(i)
        logging.info('función sacar_correos a sido ejecutada...')


def exit_set():
    print("\n\n"+ colores.GREEN + " Gracias por usar " + colores.RED + "Suite GSLT " + colores.ENDC)
    sys.exit()


def clean():
    if Sistea_Op() == "posix":
        os.system("clear")
    if Sistea_Op() == "windows":
        os.system("cls")


def Limpiar(graphic):
    if graphic == "1":
        if Sistea_Op() == "posix":
            os.system("clear")
            Azar_menu()
            print(colores.PURPLE+'\n\n\t\t\t\tCreado por: Gustavo De Jesús y Luis Trejo '+colores.ENDC)
            print(colores.BLUE+'\n\t\t\t\tFacultad de Ciencias Fisico Matematicas'+colores.ENDC)
            print(colores.BLUE+'\t\t\t\tUniversidad Autonoma de Nuevo Leon'+colores.ENDC)
            print('''\n{0}
                                    #####   #### ##### ##     ##
                                     #  #  #   #  #  #  ##   ##
                                     #    #       #     ##   ##
                                     ###  #       ###   # # # #
                                     #    #       #     # # # #
                                     #     #   #  #     #  #  #
                                    ###     ###  ###   ### # ### '''.format(colores.GREEN))
        if Sistea_Op() == "windows":
            os.system("cls")
            Azar_menu()
            print(colores.PURPLE+'\n\n\n\t\t\t\t Creado por: Gustavo De Jesús y Luis Trejo '+colores.ENDC)
            print(colores.BLUE+'\n\t\t\t\t  Facultad de Ciencias Fisico Matematicas'+colores.ENDC)
            print(colores.BLUE+'\t\t\t\t    Universidad Autonoma de Nuevo Leon'+colores.ENDC)
            print('''\n{0}
                                       #####   #### ##### ##     ##
                                        #  #  #   #  #  #  ##   ##
                                        #    #       #     ##   ##
                                        ###  #       ###   # # # #
                                        #    #       #     # # # #
                                        #     #   #  #     #  #  #
                                       ###     ###  ###   ### # ### '''.format(colores.GREEN))

def IconoScr():
    print('''

    ''')
def Azar_menu():
    opcion = random.randrange(1,5)
    if opcion == 1:
        print(colores.YELLOW + r'''
             _|_|_|            _|    _|                      _|_|_|    _|_|_|  _|    _|_|_|_|_|
           _|        _|    _|      _|_|_|_|    _|_|        _|        _|        _|        _|
             _|_|    _|    _|  _|    _|      _|_|_|_|      _|  _|_|    _|_|    _|        _|
                 _|  _|    _|  _|    _|      _|            _|    _|        _|  _|        _|
           _|_|_|      _|_|_|  _|      _|_|    _|_|_|        _|_|_|  _|_|_|    _|_|_|_|  _|   '''+colores.ENDC )
    if opcion == 2:
        print(colores.GREEN + r'''
             _______          __________________ _______    _______  _______  _    _________
            (  ____ \|\     /|\__   __|__   __/(  ____ \  (  ____ \(  ____ \( \   \__   __/
            | (    \/| )   ( |   ) (     ) (   | (    \/  | (    \/| (    \/| (      ) (
            | (_____ | |   | |   | |     | |   | (__      | |      | (_____ | |      | |
            (_____  )| |   | |   | |     | |   |  __)     | | ____ (_____  )| |      | |
                  ) || |   | |   | |     | |   | (        | | \_  )      ) || |      | |
            /\____) || (___) |___) (___  | |   | (____/\  | (___) |/\____) || (____/\| |
            \_______)(_______)\_______/  )_(   (_______/  (_______)\_______)(_______/)_(   '''+colores.ENDC)
    if opcion == 3:
        print(colores.CYAN + r'''
                         ####         #               ####   #### ###   #####
                        #   #             #          #   #  #   #  #    # # #
                        ##    ## ##  ##  ###   ##   #       ##     #      #
                         ###   #  #   #   #   #  #  #        ###   #      #
                           ##  #  #   #   #   ####  #   ###    ##  #      #
                        #   #  #  #   #   #   #      #   #  #   #  #  #   #
                        ####    #### ###   ##  ###    ####  ####  #####  ###  '''+colores.ENDC)
    if opcion == 4:
        print(colores.PURPLE + r'''
            /\\ \\                /\\                /\\\\     /\\ \\  /\\    /\\\ /\\\\\\
          /\\    /\\          /\  /\\              /\    /\\ /\\    /\\/\\         /\\
           /\\      /\\  /\\    /\/\ /\   /\\     /\\         /\\      /\\         /\\
             /\\    /\\  /\\ /\\  /\\   /\   /\\  /\\           /\\    /\\         /\\
                /\\ /\\  /\\ /\\  /\\  /\\\\\ /\\ /\\   /\\\\      /\\ /\\         /\\
          /\\    /\\/\\  /\\ /\\  /\\  /\          /\\    /\ /\\    /\\/\          /\\
            /\\ \\    /\\/\\ /\\   /\\   /\\\\      /\\\\\     /\\ \\  /\\\\\\\\   /\\    '''+colores.ENDC)
    if opcion == 5:
        print(colores.GREEN + r'''
            .d8888. db    db d888888b d888888b d88888b    d888b  .d8888. db      d888888b
            88'  YP 88    88   `88'   `~~88~~' 88'       88' Y8b 88'  YP 88      `~~88~~'
            `8bo.   88    88    88       88    88ooooo   88      `8bo.   88         88
              `Y8b. 88    88    88       88    88~~~~~   88  ooo   `Y8b. 88         88
            db   8D 88b  d88   .88.      88    88.       88. ~8~ db   8D 88booo.    88
            `8888Y' ~Y8888P' Y888888P    YP    Y88888P    Y888P  `8888Y' Y88888P    YP
                                                                                          '''+ colores.ENDC)


def Menu_Opc():
    opc = ''
    while opc !=1 or opc !=2 or opc !=3 or  opc !=4 or opc !=0:
        Azar_menu()
        print(colores.CYAN+'\n\n\n\t\tBienvenido al menu este programa cuenta con multiples modos de uso '+colores.ENDC)
        print('''\n
{0}        #################################################################################{2}
                                     [1]. Web Scraping
                                     [2]. Análisis de Metadata
                                     [3]. Ossint con Hunter
                                     [4]. Conexión por Socket
                                     [0]. Para salir
{0}        #################################################################################{1}'''.format(colores.PURPLE,colores.ENDC,colores.CYAN))
        try:
            opc = int(input("{0} \n\n Elige una de las opciones: ".format(colores.CYAN)))
            if opc == 1:
                clean()
                desicion = ''
                while desicion != 'a' or desicion != 'b' or desicion != 'c' or desicion != 'd':
                    #Azar_menu()
                    print(colores.CYAN+'\n\n      ########################################################################################## '+colores.ENDC)
                    print(colores.BLUE + r'''
           .       __        _             _____
           /       |    ___  \ ___        (        ___  .___    ___  \,___, ` , __     ___.
           |       |  .'   ` |/   \        `--.  .'   ` /   \  /   ` |    \ | |'  `. .'   `
           |  /\   /  |----' |    `           |  |      |   ' |    | |    | | |    | |    |
           |,'  \,'   `.___, `___,'      \___.'   `._.' /     `.__/| |`---' / /    |  `---|
                                                                     \                \___/
                                                                                                      '''+ colores.ENDC)
                    print('\n\n a) Descargar imagenes de una web\n b) Descargar documentos pdf de una web\n c) Descargar los URL de una web \n d) Extraer correos de la web\n e) Regresar al menu principal')
                    try:
                        desicion = input('\n Ingrese una acción: ')
                        desicion = desicion.lower() #¡Por si llega a usar mayusculas por error xD!
                        if desicion == 'a':
                            try:
                                clean()
                                Azar_menu()
                                scraping = Scraping()
                                url = input('\n\n Ingrese la URL: ')
                                directorio = input(' Ingrese la carpeta donde se guardara: ')
                                if url == '' or directorio == '':
                                    print(' ingrese una URL ')
                                    clean()
                                    break
                                scraping.scrapingImages(url,directorio)
                                input('\n\n Enter para continuar...')
                                clean()
                            except ValueError:
                                clean()
                        elif desicion == 'b':
                            try:
                                clean()
                                Azar_menu()
                                scraping = Scraping()
                                url = input('\n\n Ingrese la URL: ')
                                directorio = input(' Ingrese la carpeta donde se guardara: ')
                                if url == '' or directorio == '':
                                    print(' Ingrese una URL ')
                                    clean()
                                    break
                                scraping.scrapingPDF(url, directorio)
                                input('\n\n Enter para continuar...')
                                clean()
                            except ValueError:
                                clean()
                        elif desicion == 'c':
                            try:
                                clean()
                                Azar_menu()
                                scraping = Scraping()
                                url = input('\n\n Ingrese la URL: ')
                                directorio = input(' Ingrese la carpeta donde se guardara: ')
                                if url == '' or directorio == '':
                                    print(' Ingrese una URL ')
                                    clean()
                                    break
                                txt = input(' Ingresa el nombre del archivo txt donde se guardara: ')
                                scraping.scrapingLinks(url, directorio,txt)
                                input('\n\n Enter para continuar...')
                                clean()
                            except ValueError:
                                clean()
                        elif desicion == 'd':
                            sacar_correos()
                            input('\n\n Enter para continuar...')
                            clean()
                        elif desicion == 'e':
                            break
                        else:
                            clean()
                    except ValueError:
                        print(' El dato debe ser una opcion de menu')
                        clean()
            if opc == 2:
                clean()
                #Azar_menu()
                print(colores.CYAN+'\n\n      ########################################################################################## '+colores.ENDC)
                print(colores.PURPLE + r'''
                         __   __         .                _         .
                         |    |    ___  _/_     ___    ___/   ___  _/_     ___
                         |\  /|  .'   `  |     /   `  /   |  /   `  |     /   `
                         | \/ |  |----'  |    |    | ,'   | |    |  |    |    |
                         /    /  `.___,  \__/ `.__/| `___,' `.__/|  \__/ `.__/|
                                                          `
                                                                                                  '''+ colores.ENDC)
                #print('ingrese Metadata.py -C [aquí va el directorio de las imagenes...]\n\n')
                os.system('Metadata.py -h')
                try:
                    x = input(':')
                    os.system(x)
                except AttributeError:
                    input('No se cumple con la sintaxis [Metadata.py -C], Enter para continuar...')
                    clean()

            if opc == 3:
                clean()
                #Azar_menu()
                print(colores.CYAN+'\n\n      ########################################################################################## '+colores.ENDC)
                print(colores.RED + r'''
                     .    .___  _      __  __ .     . __    _  _______ .____  .___
                    /|    /   \ |      |   |  /     / |\   |  '   /    /      /   \
                   /  \   |,_-' |      |___|  |     | | \  |      |    |__.   |__-'
                  /---'\  |     |      |   |  |     | |  \ |      |    |      |  \
                ,'      \ /     /      /   /   `._.'  |   \|      /    /----/ /   \
                                                                                                  '''+ colores.ENDC)
                x = input('\n\n Ingrese su API KEY de Hunter: ')
                hunter = PyHunter(x)
                if len(x) < 40:
                    input(' la API KEY no es correcta: ')
                    clean()
                    break

                def Busqueda(organizacion):
                    '#Cantidad de resultados esperados de la búsqueda'
                    '#El límite MENSUAL de Hunter es 50, cuidado!'
                    busqueda = 1
                    resultado = hunter.domain_search(company=organizacion, limit=busqueda, \
                                                     emails_type='personal')

                    return resultado


                def GuardarInformacion(datosEncontrados, organizacion):
                    libro = Workbook()
                    hoja = libro.create_sheet(organizacion)
                    libro.save("Hunter" + organizacion + ".xlsx")
                    hoja.cell(1, 1, "Dominio")
                    hoja.cell(1, 2, datosEncontrados['domain'])
                    hoja.cell(2, 1, "Organización")
                    hoja.cell(2, 2, datosEncontrados['organization'])
                    hoja.cell(3, 1, "Correo")
                    hoja.cell(3, 2, datosEncontrados['emails'][0]['value'])
                    hoja.cell(4, 1, "Nombre(s)")
                    hoja.cell(4, 2, datosEncontrados['emails'][0]['first_name'])
                    hoja.cell(5, 1, "Apellidos")
                    hoja.cell(5, 2, datosEncontrados['emails'][0]['last_name'])
                    libro.save("Hunter" + organizacion + ".xlsx")


                print(" Script para buscar información")
                orga = input(" Organización a investigar: ")
                if orga == '':
                    input(' Debes escribir una organizacion...')
                datosEncontrados = Busqueda(orga)
                if datosEncontrados is None:
                    exit()
                else:
                    print(datosEncontrados)#mejorar impresion...
                    GuardarInformacion(datosEncontrados, orga)
                input (" Presione enter para continuar")
                clean()

            if opc == 4:
                clean()
                desicion = ''
                while desicion != 'a' or desicion != 'b' or desicion != 'c':
                    #Azar_menu()
                    print(colores.CYAN+'\n\n      ########################################################################################## '+colores.ENDC)
                    print(colores.YELLOW + r'''
                               _____               \             .
                              (        __.    ___  |   ,   ___  _/_
                               `--.  .'   \ .'   ` |  /  .'   `  |
                                  |  |    | |      |-<   |----'  |
                             \___.'   `._.'  `._.' /  \_ `.___,  \__/
                                                                                                      '''+ colores.ENDC)
                    print(colores.RED+'''\n                                            IMPORTANTE
                 Para hacer una conexión por socket es necesario ejecutarlo en dos
                terminales en la primera se ejecutará server y en la otra el cliente'''+colores.ENDC)
                    print('\n\n a) Ejecutar conexión host server\n b) Ejecutar conexión cliente \n c) Volver al menú principal')
                    try:
                        desicion = input('\n Ingrese una accion: ')
                        desicion = desicion.lower() #¡Por si llega a usar mayusculas por error xD!
                        if desicion == 'a':
                            try:
                                clean()
                                Azar_menu()
                                ser = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
                                ser.bind(('localhost', 8050))
                                ser.listen(1)
                                cli, addr = ser.accept()
                                while True:
                                    recibido = cli.recv(1024)
                                    print('Recibo conexion de la IP: ' + str(addr[0]) + ' Puerto: ' + str(addr[1]))
                                    print(recibido)
                                    msg_toSend=('mensaje recibido desde el la mejor profe del mundo')
                                    cli.send(msg_toSend.encode('ascii'))
                                cli.close()
                            except ValueError:
                                clean()
                        elif desicion == 'b':
                            try:
                                clean()
                                Azar_menu()
                                host = 'localhost'
                                port = 8050
                                obj = socket.socket()
                                obj.connect((host, port))
                                print('Conectado al servidor')
                                while True:
                                    mens = input('Mensaje desde Cliente a Servidor >> ')
                                    obj.send(mens.encode('ascii'))
                                    recibido = obj.recv(1024)
                                    print(recibido)
                                obj.close()
                                print('Conexión cerrada')
                            except ValueError:
                                clean()
                        elif desicion == 'c':
                            break
                        else:
                            clean()
                    except ValueError:
                        print('El dato debe ser una opcion de menu')
                        clean()
            if opc == 0:
                exit_set()
            else:
                clean()
        except ValueError:
            #input('El dato debe ser un número en el menu presione enter para continuar')
            clean()
