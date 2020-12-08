from pyhunter import PyHunter
from openpyxl import Workbook

#Sustituye la siguiente API key por la tuya
x = input('la key')
hunter = PyHunter(x)


def Busqueda(organizacion):
    '#Cantidad de resultados esperados de la búsqueda'
    '#El límite MENSUAL de Hunter es 50, cuidado!'
    busqueda = 1
    resultado = hunter.domain_search(company=organizacion, limit=busqueda, \
                                     emails_type='personal')

    return resultado


def GuardarInformacion(datosEncontrados, organizacion):
    libro = Workbook()
    hoja = libro.create_sheet(organizacion)
    libro.save("Hunter" + organizacion + ".xlsx")
    hoja.cell(1, 1, "Dominio")
    hoja.cell(1, 2, datosEncontrados['domain'])
    hoja.cell(2, 1, "Organización")
    hoja.cell(2, 2, datosEncontrados['organization'])
    hoja.cell(3, 1, "Correo")
    hoja.cell(3, 2, datosEncontrados['emails'][0]['value'])
    hoja.cell(4, 1, "Nombre(s)")
    hoja.cell(4, 2, datosEncontrados['emails'][0]['first_name'])
    hoja.cell(5, 1, "Apellidos")
    hoja.cell(5, 2, datosEncontrados['emails'][0]['last_name'])
    libro.save("Hunter" + organizacion + ".xlsx")


print("Script para buscar información")
orga = input("Organización a investigar: ")
datosEncontrados = Busqueda(orga)
if datosEncontrados is None:
    exit()
else:
    print(datosEncontrados)
    print(type(datosEncontrados))
    GuardarInformacion(datosEncontrados, orga)
